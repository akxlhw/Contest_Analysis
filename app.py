"""ICPC竞赛选手标签识别 - Streamlit Web UI (优化版)"""

import io

import streamlit as st
import pandas as pd

from config import (
    DEFAULT_EC_FINAL_KEYWORDS,
    DEFAULT_GRADE_ALIASES,
    DEFAULT_WF_KEYWORDS,
    LABEL_EXCELLENT,
    LABEL_TOP,
)
from processor import process_excel

# ============================================================
# 页面配置
# ============================================================
st.set_page_config(
    page_title="ICPC竞赛选手标签识别",
    page_icon="🏆",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ============================================================
# 自定义CSS注入
# ============================================================
st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+SC:wght@300;400;500;700;900&display=swap');

    html, body, [class*="css"] {
        font-family: 'Noto Sans SC', sans-serif;
    }

    /* 隐藏顶部装饰条 */
    header[data-testid="stHeader"] {
        display: none;
    }

    /* 主区域背景 */
    .main .block-container {
        padding-top: 0rem;
        padding-bottom: 2rem;
    }

    /* ===== Hero 区域 ===== */
    .hero-container {
        background: linear-gradient(135deg, #0f172a 0%, #1e293b 50%, #334155 100%);
        padding: 3rem 2rem 2.5rem 2rem;
        margin: -1rem -4rem 2rem -4rem;
        position: relative;
        overflow: hidden;
    }
    .hero-container::before {
        content: "";
        position: absolute;
        top: -50%;
        right: -10%;
        width: 600px;
        height: 600px;
        background: radial-gradient(circle, rgba(251,191,36,0.08) 0%, transparent 70%);
        pointer-events: none;
    }
    .hero-container::after {
        content: "";
        position: absolute;
        bottom: -30%;
        left: -5%;
        width: 400px;
        height: 400px;
        background: radial-gradient(circle, rgba(99,102,241,0.06) 0%, transparent 70%);
        pointer-events: none;
    }
    .hero-title {
        font-size: 2.4rem;
        font-weight: 900;
        color: #f8fafc;
        margin-bottom: 0.5rem;
        letter-spacing: -0.02em;
        position: relative;
        z-index: 1;
    }
    .hero-subtitle {
        font-size: 1.05rem;
        color: #94a3b8;
        font-weight: 300;
        position: relative;
        z-index: 1;
    }
    .hero-badge {
        display: inline-block;
        background: rgba(251, 191, 36, 0.15);
        color: #fbbf24;
        padding: 0.25rem 0.75rem;
        border-radius: 9999px;
        font-size: 0.75rem;
        font-weight: 600;
        margin-left: 0.75rem;
        border: 1px solid rgba(251, 191, 36, 0.25);
        position: relative;
        top: -4px;
        z-index: 1;
    }

    /* ===== 卡片 ===== */
    .card {
        background: #ffffff;
        border-radius: 16px;
        padding: 1.5rem;
        box-shadow: 0 1px 3px rgba(0,0,0,0.04), 0 4px 12px rgba(0,0,0,0.03);
        border: 1px solid #f1f5f9;
        margin-bottom: 1.25rem;
        transition: all 0.2s ease;
    }
    .card:hover {
        box-shadow: 0 4px 6px rgba(0,0,0,0.04), 0 10px 20px rgba(0,0,0,0.04);
        transform: translateY(-1px);
    }
    .card-title {
        font-size: 0.875rem;
        font-weight: 600;
        color: #475569;
        margin-bottom: 1rem;
        text-transform: uppercase;
        letter-spacing: 0.05em;
    }

    /* ===== 统计卡片 ===== */
    .stat-card {
        background: #ffffff;
        border-radius: 16px;
        padding: 1.75rem 1.5rem;
        box-shadow: 0 1px 3px rgba(0,0,0,0.04), 0 4px 12px rgba(0,0,0,0.03);
        border: 1px solid #f1f5f9;
        text-align: center;
        position: relative;
        overflow: hidden;
        transition: all 0.25s ease;
    }
    .stat-card:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 16px rgba(0,0,0,0.06);
    }
    .stat-card-top {
        border-top: 4px solid #fbbf24;
    }
    .stat-card-top::before {
        content: "";
        position: absolute;
        top: 0; left: 0; right: 0; bottom: 0;
        background: linear-gradient(180deg, rgba(251,191,36,0.04) 0%, transparent 60%);
        pointer-events: none;
    }
    .stat-card-excellent {
        border-top: 4px solid #6366f1;
    }
    .stat-card-excellent::before {
        content: "";
        position: absolute;
        top: 0; left: 0; right: 0; bottom: 0;
        background: linear-gradient(180deg, rgba(99,102,241,0.04) 0%, transparent 60%);
        pointer-events: none;
    }
    .stat-card-total {
        border-top: 4px solid #10b981;
    }
    .stat-card-total::before {
        content: "";
        position: absolute;
        top: 0; left: 0; right: 0; bottom: 0;
        background: linear-gradient(180deg, rgba(16,185,129,0.04) 0%, transparent 60%);
        pointer-events: none;
    }
    .stat-icon {
        font-size: 1.75rem;
        margin-bottom: 0.5rem;
    }
    .stat-value {
        font-size: 2.5rem;
        font-weight: 900;
        color: #0f172a;
        line-height: 1;
        margin-bottom: 0.25rem;
        position: relative;
        z-index: 1;
    }
    .stat-label {
        font-size: 0.875rem;
        color: #64748b;
        font-weight: 500;
        position: relative;
        z-index: 1;
    }

    /* ===== 规则卡片 ===== */
    .rule-section {
        background: #f8fafc;
        border-radius: 12px;
        padding: 1rem 1.25rem;
        margin-bottom: 0.75rem;
        border-left: 4px solid #fbbf24;
    }
    .rule-section.excellent-rule {
        border-left-color: #6366f1;
    }
    .rule-title {
        font-weight: 700;
        color: #1e293b;
        font-size: 0.95rem;
        margin-bottom: 0.4rem;
    }
    .rule-list {
        color: #475569;
        font-size: 0.9rem;
        line-height: 1.6;
        margin: 0;
        padding-left: 1.2rem;
    }
    .rule-note {
        margin-top: 0.75rem;
        padding-top: 0.75rem;
        border-top: 1px dashed #cbd5e1;
        color: #64748b;
        font-size: 0.85rem;
    }

    /* ===== 标签药丸 ===== */
    .pill {
        display: inline-block;
        padding: 0.35rem 0.9rem;
        border-radius: 9999px;
        font-size: 0.8rem;
        font-weight: 600;
    }
    .pill-top {
        background: linear-gradient(135deg, #fef3c7, #fde68a);
        color: #92400e;
        border: 1px solid #fbbf24;
    }
    .pill-excellent {
        background: linear-gradient(135deg, #e0e7ff, #c7d2fe);
        color: #3730a3;
        border: 1px solid #818cf8;
    }

    /* ===== 文件信息条 ===== */
    .file-info {
        background: linear-gradient(90deg, #f0f9ff, #e0f2fe);
        border: 1px solid #bae6fd;
        border-radius: 12px;
        padding: 1rem 1.5rem;
        display: flex;
        align-items: center;
        gap: 1rem;
        color: #0369a1;
        font-weight: 500;
        font-size: 0.95rem;
    }
    .file-info-icon {
        font-size: 1.5rem;
    }

    /* ===== 空状态 ===== */
    .empty-state {
        text-align: center;
        padding: 4rem 2rem;
        color: #94a3b8;
    }
    .empty-state-icon {
        font-size: 4rem;
        margin-bottom: 1rem;
        opacity: 0.5;
    }
    .empty-state-text {
        font-size: 1.1rem;
        font-weight: 500;
    }

    /* ===== 分割线美化 ===== */
    hr.custom-hr {
        border: none;
        height: 1px;
        background: linear-gradient(90deg, transparent, #e2e8f0, transparent);
        margin: 1.5rem 0;
    }

    /* ===== 下载按钮区域 ===== */
    .download-area {
        background: #f8fafc;
        border-radius: 12px;
        padding: 1.25rem;
        border: 1px dashed #cbd5e1;
        margin-top: 1rem;
    }

    /* ===== Streamlit 原生组件微调 ===== */
    div[data-testid="stFileUploader"] {
        background: #ffffff;
        border-radius: 16px;
        padding: 1.5rem;
        border: 2px dashed #cbd5e1;
        transition: all 0.2s ease;
    }
    div[data-testid="stFileUploader"]:hover {
        border-color: #6366f1;
        background: #fafafa;
    }
    div[data-testid="stFileUploader"] > label {
        font-weight: 600;
        color: #334155;
    }

    .stButton > button {
        border-radius: 10px !important;
        font-weight: 600 !important;
        padding: 0.5rem 1.5rem !important;
        transition: all 0.2s ease !important;
    }
    .stButton > button:hover {
        transform: translateY(-1px);
        box-shadow: 0 4px 12px rgba(0,0,0,0.1) !important;
    }

    /* sidebar */
    section[data-testid="stSidebar"] {
        background: #f8fafc;
    }
    section[data-testid="stSidebar"] .block-container {
        padding-top: 2rem;
    }

    /* dataframe */
    .dataframe-container {
        border-radius: 12px;
        overflow: hidden;
        border: 1px solid #e2e8f0;
        box-shadow: 0 1px 3px rgba(0,0,0,0.02);
    }

    /* 动画 */
    @keyframes fadeInUp {
        from { opacity: 0; transform: translateY(12px); }
        to { opacity: 1; transform: translateY(0); }
    }
    .animate-in {
        animation: fadeInUp 0.5s ease forwards;
    }
    .delay-1 { animation-delay: 0.05s; opacity: 0; }
    .delay-2 { animation-delay: 0.1s; opacity: 0; }
    .delay-3 { animation-delay: 0.15s; opacity: 0; }
    .delay-4 { animation-delay: 0.2s; opacity: 0; }

    /* 关键词标签 */
    .kw-tag {
        display: inline-flex;
        align-items: center;
        gap: 0.35rem;
        background: #f1f5f9;
        color: #475569;
        padding: 0.3rem 0.7rem;
        border-radius: 8px;
        font-size: 0.8rem;
        font-weight: 500;
        margin: 0.15rem;
        border: 1px solid #e2e8f0;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# ============================================================
# Hero 区域
# ============================================================
st.markdown(
    f"""
    <div class="hero-container">
        <div class="hero-title">
            🏆 ICPC 竞赛选手标签识别
            <span class="hero-badge">智能标注</span>
        </div>
        <div class="hero-subtitle">
            自动识别并标注 EC-Final / World-Final 竞赛选手 achievement 等级
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

# ============================================================
# 初始化session_state
# ============================================================
if "ec_keywords" not in st.session_state:
    st.session_state.ec_keywords = list(DEFAULT_EC_FINAL_KEYWORDS)
if "wf_keywords" not in st.session_state:
    st.session_state.wf_keywords = list(DEFAULT_WF_KEYWORDS)
if "grade_aliases" not in st.session_state:
    st.session_state.grade_aliases = {k: list(v) for k, v in DEFAULT_GRADE_ALIASES.items()}

# ============================================================
# 侧边栏: 配置
# ============================================================
with st.sidebar:
    st.markdown(
        """
        <div style="margin-bottom:1.5rem;">
            <div style="font-size:1.1rem;font-weight:700;color:#1e293b;margin-bottom:0.25rem;">⚙️ 配置面板</div>
            <div style="font-size:0.8rem;color:#94a3b8;">调整处理参数</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    header_row = st.number_input("表头行号", min_value=1, max_value=10, value=1, step=1)

    st.markdown("<hr class='custom-hr'>", unsafe_allow_html=True)

    st.markdown(
        """
        <div style="margin-bottom:1rem;">
            <div style="font-size:0.95rem;font-weight:600;color:#475569;">📋 Excel 格式说明</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.markdown(
        """
        <div style="background:#ffffff;border-radius:10px;padding:0.9rem;border:1px solid #e2e8f0;font-size:0.8rem;color:#64748b;line-height:1.7;">
        <b style="color:#334155;">基本信息列</b><br>
        A列: 序号 &nbsp;|&nbsp; B列: 姓名<br>
        C列: 学号 &nbsp;|&nbsp; D列: 学院<br><br>
        <b style="color:#334155;">竞赛记录列 (5组)</b><br>
        CS~CV : 竞赛 1<br>
        CW~CZ : 竞赛 2<br>
        DA~DD : 竞赛 3<br>
        DE~DH : 竞赛 4<br>
        DI~DL : 竞赛 5<br><br>
        <span style="color:#0ea5e9;">每组4列：名称、时间、级别、等级</span>
        </div>
        """,
        unsafe_allow_html=True,
    )

# ============================================================
# 主体内容区
# ============================================================

# ----- 文件上传区块 -----
st.markdown(
    """
    <div style="margin-bottom:0.75rem;">
        <span style="font-size:1.05rem;font-weight:700;color:#1e293b;">📤 上传数据</span>
        <span style="font-size:0.8rem;color:#94a3b8;margin-left:0.5rem;">支持 .xlsx / .xls 格式</span>
    </div>
    """,
    unsafe_allow_html=True,
)

uploaded_file = st.file_uploader("", type=["xlsx", "xls"], label_visibility="collapsed")

if not uploaded_file:
    st.markdown(
        """
        <div class="empty-state animate-in">
            <div class="empty-state-icon">📂</div>
            <div class="empty-state-text">请上传包含学生竞赛记录的 Excel 文件</div>
            <div style="font-size:0.85rem;margin-top:0.5rem;">文件将仅在内存中处理，不会上传至服务器</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.stop()

# 读取基本信息
import openpyxl

wb = openpyxl.load_workbook(uploaded_file)
ws = wb.active

st.markdown(
    f"""
    <div class="file-info animate-in">
        <span class="file-info-icon">📊</span>
        <div>
            <div style="font-weight:700;color:#0c4a6e;">{uploaded_file.name}</div>
            <div style="font-size:0.8rem;color:#38bdf8;margin-top:0.15rem;">
                {ws.max_row} 行 &nbsp;·&nbsp; {ws.max_column} 列 &nbsp;·&nbsp; Sheet: {ws.title}
            </div>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)
del wb

st.markdown("<hr class='custom-hr'>", unsafe_allow_html=True)

# ============================================================
# 匹配规则说明
# ============================================================
with st.expander("📖 匹配规则说明", expanded=False):
    st.markdown(
        f"""
        <div style="display:flex;flex-wrap:wrap;gap:1rem;">
            <div class="rule-section" style="flex:1;min-width:280px;">
                <div class="rule-title">🥇 {LABEL_TOP}</div>
                <ul class="rule-list">
                    <li>EC-Final 金奖</li>
                    <li>World-Final 金奖 / 银奖 / 铜奖</li>
                </ul>
            </div>
            <div class="rule-section excellent-rule" style="flex:1;min-width:280px;">
                <div class="rule-title">🥈 {LABEL_EXCELLENT}</div>
                <ul class="rule-list">
                    <li>EC-Final 银奖 / 铜奖</li>
                    <li>World-Final 参赛但未获金银铜奖</li>
                </ul>
            </div>
        </div>
        <div class="rule-note">
            💡 <b>优先级规则</b>：同时满足两个等级时取最高等级（{LABEL_TOP} 优先）
        </div>
        """,
        unsafe_allow_html=True,
    )

# ============================================================
# 关键词与别名管理
# ============================================================
with st.expander("🔧 关键词与别名管理", expanded=False):
    st.markdown('<div class="card-title">竞赛名称关键词</div>', unsafe_allow_html=True)

    col_ec, col_wf = st.columns(2)

    with col_ec:
        st.markdown("**EC-Final 关键词**")
        for i, kw in enumerate(st.session_state.ec_keywords):
            c1, c2 = st.columns([4, 1])
            c1.markdown(f'<span class="kw-tag">{kw}</span>', unsafe_allow_html=True)
            if c2.button("删除", key=f"del_ec_{i}"):
                st.session_state.ec_keywords.pop(i)
                st.rerun()

        c1, c2 = st.columns([3, 1])
        new_ec = c1.text_input("添加EC-Final关键词", key="new_ec_input", label_visibility="collapsed", placeholder="输入新关键词...")
        if c2.button("➕ 添加", key="add_ec_btn") and new_ec.strip():
            if new_ec.strip() not in st.session_state.ec_keywords:
                st.session_state.ec_keywords.append(new_ec.strip())
                st.rerun()

    with col_wf:
        st.markdown("**WF 关键词**")
        for i, kw in enumerate(st.session_state.wf_keywords):
            c1, c2 = st.columns([4, 1])
            c1.markdown(f'<span class="kw-tag">{kw}</span>', unsafe_allow_html=True)
            if c2.button("删除", key=f"del_wf_{i}"):
                st.session_state.wf_keywords.pop(i)
                st.rerun()

        c1, c2 = st.columns([3, 1])
        new_wf = c1.text_input("添加WF关键词", key="new_wf_input", label_visibility="collapsed", placeholder="输入新关键词...")
        if c2.button("➕ 添加", key="add_wf_btn") and new_wf.strip():
            if new_wf.strip() not in st.session_state.wf_keywords:
                st.session_state.wf_keywords.append(new_wf.strip())
                st.rerun()

    st.markdown("<hr class='custom-hr'>", unsafe_allow_html=True)
    st.markdown('<div class="card-title">等级别名管理</div>', unsafe_allow_html=True)

    alias_cols = st.columns(3)
    for idx, standard_grade in enumerate(["金奖", "银奖", "铜奖"]):
        with alias_cols[idx]:
            st.markdown(f"**{standard_grade} 别名**")
            aliases = st.session_state.grade_aliases[standard_grade]
            for i, alias in enumerate(aliases):
                c1, c2 = st.columns([3, 1])
                c1.markdown(f"<span class='kw-tag'>{alias}</span>", unsafe_allow_html=True)
                if c2.button("×", key=f"del_grade_{standard_grade}_{i}"):
                    st.session_state.grade_aliases[standard_grade].pop(i)
                    st.rerun()

            new_alias = st.text_input(f"添加", key=f"new_grade_{standard_grade}", label_visibility="collapsed", placeholder="新别名...")
            if st.button(f"添加", key=f"add_grade_{standard_grade}_btn") and new_alias.strip():
                if new_alias.strip() not in st.session_state.grade_aliases[standard_grade]:
                    st.session_state.grade_aliases[standard_grade].append(new_alias.strip())
                    st.rerun()

    st.markdown("<div style='margin-top:1rem;'></div>", unsafe_allow_html=True)
    if st.button("🔄 重置为默认", key="reset_keywords", type="secondary"):
        st.session_state.ec_keywords = list(DEFAULT_EC_FINAL_KEYWORDS)
        st.session_state.wf_keywords = list(DEFAULT_WF_KEYWORDS)
        st.session_state.grade_aliases = {k: list(v) for k, v in DEFAULT_GRADE_ALIASES.items()}
        st.rerun()

st.markdown("<hr class='custom-hr'>", unsafe_allow_html=True)

# ============================================================
# 操作按钮
# ============================================================
st.markdown(
    """
    <div style="margin-bottom:0.75rem;">
        <span style="font-size:1.05rem;font-weight:700;color:#1e293b;">🚀 开始处理</span>
    </div>
    """,
    unsafe_allow_html=True,
)

col_dry, col_run, _ = st.columns([1, 1, 3])

result = None
with col_dry:
    dry_run = st.button("👁️ Dry Run 预览", type="secondary", use_container_width=True)
with col_run:
    run_process = st.button("▶️ 开始处理", type="primary", use_container_width=True)

if dry_run or run_process:
    with st.spinner("正在分析竞赛记录，请稍候..."):
        file_buffer = io.BytesIO(uploaded_file.getvalue())
        labeled_bytes, qualified_df, stats = process_excel(
            file_buffer,
            ec_keywords=st.session_state.ec_keywords,
            wf_keywords=st.session_state.wf_keywords,
            grade_aliases=st.session_state.grade_aliases,
            header_row=header_row,
        )
        result = (labeled_bytes, qualified_df, stats)

    if dry_run:
        st.session_state.dry_result = result
        st.session_state.process_result = None
        st.toast("✅ 预览模式：标签未实际写入文件", icon="👁️")
    else:
        st.session_state.process_result = result
        st.session_state.dry_result = None
        st.toast("✅ 处理完成！可下载结果文件", icon="🎉")

# ============================================================
# 处理结果
# ============================================================
active_result = st.session_state.get("process_result") or st.session_state.get("dry_result")

if active_result:
    labeled_bytes, qualified_df, stats = active_result

    st.markdown("<hr class='custom-hr'>", unsafe_allow_html=True)

    # 模式标签
    is_dry = st.session_state.get("dry_result") is not None
    mode_badge = (
        '<span class="pill pill-excellent" style="margin-left:0.5rem;">预览模式</span>'
        if is_dry
        else '<span class="pill pill-top" style="margin-left:0.5rem;">已处理</span>'
    )
    st.markdown(
        f"""
        <div style="margin-bottom:1rem;">
            <span style="font-size:1.05rem;font-weight:700;color:#1e293b;">📈 处理结果</span>
            {mode_badge}
        </div>
        """,
        unsafe_allow_html=True,
    )

    # 统计卡片
    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown(
            f"""
            <div class="stat-card stat-card-top animate-in delay-1">
                <div class="stat-icon">🥇</div>
                <div class="stat-value">{stats.get(LABEL_TOP, 0)}</div>
                <div class="stat-label">{LABEL_TOP}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with c2:
        st.markdown(
            f"""
            <div class="stat-card stat-card-excellent animate-in delay-2">
                <div class="stat-icon">🥈</div>
                <div class="stat-value">{stats.get(LABEL_EXCELLENT, 0)}</div>
                <div class="stat-label">{LABEL_EXCELLENT}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with c3:
        st.markdown(
            f"""
            <div class="stat-card stat-card-total animate-in delay-3">
                <div class="stat-icon">🏆</div>
                <div class="stat-value">{stats.get("总计", 0)}</div>
                <div class="stat-label">合格选手总计</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.markdown("<div style='height:1rem;'></div>", unsafe_allow_html=True)

    if not qualified_df.empty:
        # 筛选
        filter_options = ["全部", LABEL_TOP, LABEL_EXCELLENT]
        filter_label = st.selectbox("按标签筛选", filter_options, index=0)
        if filter_label != "全部":
            display_df = qualified_df[qualified_df["标签"] == filter_label]
        else:
            display_df = qualified_df

        st.markdown(
            f"""
            <div style="margin-bottom:0.5rem;font-size:0.85rem;color:#64748b;">
                共显示 <b>{len(display_df)}</b> 条记录
            </div>
            """,
            unsafe_allow_html=True,
        )

        st.dataframe(
            display_df,
            use_container_width=True,
            hide_index=True,
            column_config={
                "标签": st.column_config.TextColumn(
                    "标签",
                    help="选手 achievement 等级",
                ),
                "达标原因": st.column_config.TextColumn(
                    "达标原因",
                    help="满足条件的竞赛及奖项",
                ),
            },
        )

        # 下载按钮
        if st.session_state.get("process_result") is not None:
            st.markdown(
                """
                <div class="download-area animate-in delay-4">
                    <div style="font-weight:600;color:#334155;margin-bottom:0.75rem;">📥 下载结果</div>
                """,
                unsafe_allow_html=True,
            )
            dl1, dl2 = st.columns(2)
            with dl1:
                st.download_button(
                    "📄 下载标签文件",
                    labeled_bytes,
                    "students_labeled.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            with dl2:
                qualified_bytes = io.BytesIO()
                with pd.ExcelWriter(qualified_bytes, engine="openpyxl") as writer:
                    display_df.to_excel(writer, index=False, sheet_name="合格名单")
                qualified_bytes.seek(0)
                st.download_button(
                    "📋 下载合格名单",
                    qualified_bytes.getvalue(),
                    "qualified_students.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            st.markdown("</div>", unsafe_allow_html=True)
    else:
        st.warning("未识别到符合条件的选手")
