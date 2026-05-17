"""生成模拟Excel测试数据，覆盖所有33个场景"""

import openpyxl
from openpyxl.utils import get_column_letter

# 5组竞赛列 (名称, 时间, 级别, 等级)，1-based
COMP_COLS = [
    (97, 98, 99, 100),
    (101, 102, 103, 104),
    (105, 106, 107, 108),
    (109, 110, 111, 112),
    (113, 114, 115, 116),
]

# 场景定义: (姓名, [(竞赛名, 时间, 级别, 等级), ...], 预期标签)
# None表示空, 竞赛填在第1个空slot
SCENARIOS = [
    # --- 基本场景 ---
    ("张1_EC金", [("ICPC EC-Final", "2023", "国际", "金奖")], "顶尖竞赛选手"),
    ("张2_EC银", [("EC-Final", "2023", "国际", "银奖")], "卓越竞赛选手"),
    ("张3_EC铜", [("ec final", "2023", "国际", "铜奖")], "卓越竞赛选手"),
    ("张4_WF金", [("ICPC World-Final", "2023", "国际", "金奖")], "顶尖竞赛选手"),
    ("张5_WF银", [("World Final", "2023", "国际", "银奖")], "顶尖竞赛选手"),
    ("张6_WF铜", [("worldfinal", "2023", "国际", "铜奖")], "顶尖竞赛选手"),
    ("张7_WF无奖空等级", [("WF", "2023", "国际", "")], "卓越竞赛选手"),
    ("张8_WF无奖None等级", [("ICPC WF Dhaka", "2023", "国际", None)], "卓越竞赛选手"),
    ("张9_WF参赛奖", [("wf", "2023", "国际", "参赛奖")], "卓越竞赛选手"),

    # --- 组合场景 ---
    ("张10_EC金WF无奖", [("ICPC EC-Final", "2023", "国际", "金奖"), ("ICPC WF", "2023", "国际", None)], "顶尖竞赛选手"),
    ("张11_EC铜WF铜", [("EC-Final", "2023", "国际", "铜奖"), ("World Final", "2023", "国际", "铜奖")], "顶尖竞赛选手"),

    # --- 边界场景 ---
    ("张12_无关竞赛", [("蓝桥杯", "2023", "国内", "一等奖")], None),
    ("张13_全空", [], None),
    ("张14_ECO误匹配", [("ECO-Final Contest", "2023", "国内", "金奖")], None),
    ("张15_SWIFT误匹配", [("SWIFT Programming", "2023", "国内", "金奖")], None),
    ("张16_workflow误匹配", [("workflow大赛", "2023", "国内", "金奖")], None),

    # --- 名称变体 ---
    ("张17_EC点Final", [("ICPC EC.Final 2023", "2023", "国际", "铜奖")], "卓越竞赛选手"),
    ("张18_ECFinal", [("ECFinal 2023", "2023", "国际", "银奖")], "卓越竞赛选手"),
    ("张19_EC_Finals", [("ICPC EC Finals", "2023", "国际", "铜奖")], "卓越竞赛选手"),
    ("张20_World_Finals", [("ICPC World Finals 2023", "2023", "国际", "金奖")], "顶尖竞赛选手"),
    ("张21_WF点", [("W.F.", "2023", "国际", None)], "卓越竞赛选手"),
    ("张22_中文世界总决赛", [("ICPC世界总决赛", "2023", "国际", "金奖")], "顶尖竞赛选手"),

    # --- Slot 3/4/5 测试 ---
    ("张23_slot3", [(None, None, None, None), (None, None, None, None), ("EC-Final", "2023", "国际", "金奖")], "顶尖竞赛选手"),
    ("张24_slot4", [(None, None, None, None), (None, None, None, None), (None, None, None, None), ("World Final", "2023", "国际", "银奖")], "顶尖竞赛选手"),
    ("张25_slot5", [(None, None, None, None), (None, None, None, None), (None, None, None, None), (None, None, None, None), ("EC-Final", "2023", "国际", "铜奖")], "卓越竞赛选手"),

    # --- 大小写混合 ---
    ("张26_大小写", [("Ec-fInAl", "2023", "国际", "银奖")], "卓越竞赛选手"),

    # --- 等级变体 ---
    ("张27_金牌", [("EC-Final", "2023", "国际", "金牌")], "顶尖竞赛选手"),
    ("张28_第一名", [("WF", "2023", "国际", "第一名")], "顶尖竞赛选手"),
    ("张29_一等奖", [("EC-Final", "2023", "国际", "一等奖")], "顶尖竞赛选手"),
    ("张30_亚军", [("EC-Final", "2023", "国际", "亚军")], "卓越竞赛选手"),
    ("张31_三等奖", [("EC-Final", "2023", "国际", "三等奖")], "卓越竞赛选手"),
    ("张32_bronze", [("WF", "2023", "国际", "bronze")], "顶尖竞赛选手"),
    ("张33_2nd", [("WF", "2023", "国际", "2nd")], "顶尖竞赛选手"),
    ("张34_季军", [("EC-Final", "2023", "国际", "季军")], "卓越竞赛选手"),
]


def generate_mock_excel(filepath="mock_students.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "学生名单"

    # 写表头
    headers = ["序号", "姓名", "学号", "学院"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)

    # 竞赛列表头
    for slot_idx, (name_col, time_col, level_col, grade_col) in enumerate(COMP_COLS, 1):
        ws.cell(row=1, column=name_col, value=f"竞赛名称{slot_idx}")
        ws.cell(row=1, column=time_col, value=f"竞赛时间{slot_idx}")
        ws.cell(row=1, column=level_col, value=f"竞赛级别{slot_idx}")
        ws.cell(row=1, column=grade_col, value=f"获奖等级{slot_idx}")

    # DL列后加备注列(列117)测试动态定位
    ws.cell(row=1, column=117, value="备注")

    # 写数据
    for row_idx, (name, competitions, _expected) in enumerate(SCENARIOS, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=f"2023{row_idx - 1:04d}")
        ws.cell(row=row_idx, column=4, value="计算机学院")

        # 填写竞赛数据
        for comp_idx, (comp_name, comp_time, comp_level, comp_grade) in enumerate(competitions):
            if comp_idx >= 5:
                break
            name_col, time_col, level_col, grade_col = COMP_COLS[comp_idx]
            if comp_name is not None:
                ws.cell(row=row_idx, column=name_col, value=comp_name)
            if comp_time is not None:
                ws.cell(row=row_idx, column=time_col, value=comp_time)
            if comp_level is not None:
                ws.cell(row=row_idx, column=level_col, value=comp_level)
            if comp_grade is not None:
                ws.cell(row=row_idx, column=grade_col, value=comp_grade)

    wb.save(filepath)
    print(f"已生成模拟数据: {filepath} ({len(SCENARIOS)}条记录)")


if __name__ == "__main__":
    generate_mock_excel()
