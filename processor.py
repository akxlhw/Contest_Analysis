"""ICPC竞赛选手标签识别 - 核心处理模块（纯逻辑，不依赖UI）"""

import io

import openpyxl

from config import (
    COMPETITION_COLUMNS,
    DEFAULT_EC_FINAL_KEYWORDS,
    DEFAULT_GRADE_ALIASES,
    DEFAULT_WF_KEYWORDS,
    GOLD_GRADES,
    LABEL_EXCELLENT,
    LABEL_TOP,
    SILVER_GRADES,
    STUDENT_INFO_COLUMNS,
    BRONZE_GRADES,
    match_competition_name,
    normalize_grade,
)


def extract_competitions(ws, row):
    """提取一行中5组竞赛数据

    Args:
        ws: openpyxl worksheet
        row: 行号(1-based)

    Returns:
        list[dict]: 5个竞赛记录，每个含name/date/level/grade
    """
    competitions = []
    for name_col, time_col, level_col, grade_col in COMPETITION_COLUMNS:
        competitions.append({
            "name": ws.cell(row=row, column=name_col).value,
            "date": ws.cell(row=row, column=time_col).value,
            "level": ws.cell(row=row, column=level_col).value,
            "grade": ws.cell(row=row, column=grade_col).value,
        })
    return competitions


def classify_student(competitions, ec_keywords=None, wf_keywords=None, grade_aliases=None):
    """根据竞赛记录判断学生标签

    Args:
        competitions: extract_competitions返回的5组竞赛记录
        ec_keywords: EC-Final关键词列表
        wf_keywords: WF关键词列表
        grade_aliases: 等级别名映射表

    Returns:
        LABEL_TOP / LABEL_EXCELLENT / None
    """
    ec_kw = ec_keywords or DEFAULT_EC_FINAL_KEYWORDS
    wf_kw = wf_keywords or DEFAULT_WF_KEYWORDS
    ga = grade_aliases or DEFAULT_GRADE_ALIASES

    has_top = False
    has_excellent = False

    for comp in competitions:
        comp_type = match_competition_name(comp["name"], ec_kw, wf_kw)
        if comp_type is None:
            continue

        normalized = normalize_grade(comp["grade"], ga)

        if comp_type == "ec_final":
            if normalized in GOLD_GRADES:
                has_top = True
            elif normalized in SILVER_GRADES or normalized in BRONZE_GRADES:
                has_excellent = True
        elif comp_type == "wf":
            if normalized in GOLD_GRADES or normalized in SILVER_GRADES or normalized in BRONZE_GRADES:
                has_top = True
            else:
                # WF参赛但无金银铜奖牌
                has_excellent = True

    if has_top:
        return LABEL_TOP
    if has_excellent:
        return LABEL_EXCELLENT
    return None


def get_qualifying_reasons(competitions, ec_keywords=None, wf_keywords=None, grade_aliases=None):
    """获取达标原因的人类可读描述

    Args:
        competitions: 5组竞赛记录
        ec_keywords: EC-Final关键词列表
        wf_keywords: WF关键词列表
        grade_aliases: 等级别名映射表

    Returns:
        str: 如 "EC-Final金奖"、"WF银奖 + EC-Final铜奖"
    """
    ec_kw = ec_keywords or DEFAULT_EC_FINAL_KEYWORDS
    wf_kw = wf_keywords or DEFAULT_WF_KEYWORDS
    ga = grade_aliases or DEFAULT_GRADE_ALIASES

    reasons = []
    for comp in competitions:
        comp_type = match_competition_name(comp["name"], ec_kw, wf_kw)
        if comp_type is None:
            continue

        normalized = normalize_grade(comp["grade"], ga)

        if comp_type == "ec_final":
            if normalized in GOLD_GRADES:
                reasons.append("EC-Final金奖")
            elif normalized in SILVER_GRADES:
                reasons.append("EC-Final银奖")
            elif normalized in BRONZE_GRADES:
                reasons.append("EC-Final铜奖")
        elif comp_type == "wf":
            if normalized in GOLD_GRADES:
                reasons.append("WF金奖")
            elif normalized in SILVER_GRADES:
                reasons.append("WF银奖")
            elif normalized in BRONZE_GRADES:
                reasons.append("WF铜奖")
            else:
                reasons.append("WF参赛")

    return " + ".join(reasons) if reasons else ""


def find_label_column(ws):
    """定位标签写入列: 从DL之后找第一个空列

    Args:
        ws: openpyxl worksheet

    Returns:
        int: 列号(1-based)
    """
    # DL是第116列，从117开始找
    start_col = 117
    row = 1  # 检查表头行
    col = start_col
    while ws.cell(row=row, column=col).value is not None:
        col += 1
    return col


def process_excel(input_buffer, ec_keywords=None, wf_keywords=None, grade_aliases=None,
                  header_row=1):
    """主处理函数: 读取Excel，识别标签，生成结果

    Args:
        input_buffer: 上传的Excel文件字节流
        ec_keywords: EC-Final关键词列表
        wf_keywords: WF关键词列表
        grade_aliases: 等级别名映射表
        header_row: 表头行号(1-based)

    Returns:
        tuple: (labeled_bytes, qualified_df, stats_dict)
            - labeled_bytes: 添加了标签列的Excel字节流
            - qualified_df: 合格学生DataFrame（供UI展示）
            - stats_dict: {"顶尖": N, "卓越": M, "总计": K}
    """
    import pandas as pd

    ec_kw = ec_keywords or DEFAULT_EC_FINAL_KEYWORDS
    wf_kw = wf_keywords or DEFAULT_WF_KEYWORDS
    ga = grade_aliases or DEFAULT_GRADE_ALIASES

    wb = openpyxl.load_workbook(input_buffer)
    ws = wb.active

    # 定位标签列
    label_col = find_label_column(ws)
    ws.cell(row=header_row, column=label_col, value="竞赛标签")

    # 统计
    stats = {LABEL_TOP: 0, LABEL_EXCELLENT: 0}
    qualified_rows = []

    data_start = header_row + 1
    max_row = ws.max_row

    for row in range(data_start, max_row + 1):
        # 读取学生基本信息
        seq = ws.cell(row=row, column=STUDENT_INFO_COLUMNS["序号"]).value
        name = ws.cell(row=row, column=STUDENT_INFO_COLUMNS["姓名"]).value
        sid = ws.cell(row=row, column=STUDENT_INFO_COLUMNS["学号"]).value
        college = ws.cell(row=row, column=STUDENT_INFO_COLUMNS["学院"]).value

        # 提取竞赛数据
        competitions = extract_competitions(ws, row)

        # 分类
        label = classify_student(competitions, ec_kw, wf_kw, ga)

        if label:
            ws.cell(row=row, column=label_col, value=label)
            stats[label] = stats.get(label, 0) + 1

            # 获取达标原因
            reasons = get_qualifying_reasons(competitions, ec_kw, wf_kw, ga)

            # 获取匹配到的竞赛名称和等级(原始值)
            matched_comps = []
            matched_grades = []
            for comp in competitions:
                comp_type = match_competition_name(comp["name"], ec_kw, wf_kw)
                if comp_type is not None:
                    matched_comps.append(str(comp["name"]) if comp["name"] else "")
                    matched_grades.append(str(comp["grade"]) if comp["grade"] else "(无奖牌)")

            qualified_rows.append({
                "序号": seq,
                "姓名": name,
                "学号": sid,
                "学院": college,
                "标签": label,
                "达标原因": reasons,
                "竞赛名称": ", ".join(matched_comps),
                "获奖等级": ", ".join(matched_grades),
            })
        else:
            ws.cell(row=row, column=label_col, value="")

    # 生成labeled_bytes
    labeled_buffer = io.BytesIO()
    wb.save(labeled_buffer)
    labeled_buffer.seek(0)
    labeled_bytes = labeled_buffer.getvalue()

    # 生成qualified_df
    qualified_df = pd.DataFrame(qualified_rows) if qualified_rows else pd.DataFrame(
        columns=["序号", "姓名", "学号", "学院", "标签", "达标原因", "竞赛名称", "获奖等级"]
    )

    stats["总计"] = sum(stats.values())

    return labeled_bytes, qualified_df, stats
